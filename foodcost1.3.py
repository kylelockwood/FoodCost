#! python3

# TODO tkinter GUI
# TODO django interface
# TODO get current prices from vendor database
# TODO Encryption
# TODO Merge recipes/ingredients from another store

import sys
import re
import os.path
import time
import openpyxl
import json
import uuid
from string import capwords
import mysql.connector
from mysql.connector import Error
import pandas as pd
import tkinter as tk
from tkinter import messagebox

# TODO add opt cats to add_ingredient and add_recipe
    # See about removing the xl variable and using the _json dict from Ingredients and Recipes

# Globals assigned in Store.__init__()
LABOR = 0
MARKUP = 0

class Food_Cost():
    """Application for calculating the cost of a recipe from cost and quantity of ingredients and labor"""
    def __init__(self, args=None):
        self.store = None
        self._login(args)
        self.xl = { "ing_req_cats": ["name", "unit", "quantity", "price"],
                    "ing_opt_cats": ["vendor", "notes", "calories", "servings"],
                    "rec_req_cats": ["name", "ingredients", "quantites", "preptime"],
                    "rec_opt_cats": ["notes", "yield"]}
        
        self.main_menu()

    def _login(self, args=None):
        """Login to store in stores.db"""
        # TODO encript stores.json

        # Window
        self.login_window = tk.Tk()
        self.login_window.title('Food Cost')
        self.login_window.geometry('300x160')
        
        # Title
        title_label = tk.Label(self.login_window, text='Store Login')
        title_label.grid(column=1, row=0, columnspan=5, pady=15)
        
        # User input
        user_label = tk.Label(self.login_window, text='Email : ')
        user_label.grid(row=2, column=1, padx=10)
        self.user_box = tk.Entry(self.login_window, width=30)
        self.user_box.grid(column=2, row=2, columnspan=4, pady=5, sticky='W')
        if args:
            self.user_box.insert(0, args[1])
        
        # Pass input
        pass_label = tk.Label(self.login_window, text='Password : ')
        pass_label.grid(row=3, column=1, padx=10)
        self.pass_box = tk.Entry(self.login_window, show='●', width=30)
        self.pass_box.grid(column=2, row=3, columnspan=4, pady=5, sticky='W')
        if args:
            self.pass_box.insert(0, args[2])

        # Buttons
        login_button = tk.Button(self.login_window, text='Login', command=self.__check_login__, width=10)
        login_button.grid(column=2, row=4, pady=10, padx=5)
        new_button = tk.Button(self.login_window, text='New Store', command=self._create_store, width=10)
        new_button.grid(column=4, row=4, pady=10, padx=5)
        
        self.login_window.mainloop()
        
    def __check_login__(self):
        user = self.user_box.get()
        pw = self.pass_box.get()
        stores = __load_json__('stores.json')
        for store in stores:
            if stores[store]['email'] == user and stores[store]['pass'] == pw:
                messagebox.showinfo('Login Success', 'Login Success.')
                self.login_window.destroy()
                self.store = Store(__load_json__(stores[store]['db_file']))
                return
        messagebox.showinfo('Login Error', 'Incorrect email and/or password.')

    def _create_store(self, update=None):
        self.login_window.destroy()
        # TODO DRY

        # Window
        self.ns_window = tk.Tk()
        self.ns_window.title('Food Cost')
        self.ns_window.geometry('340x340')
        
        # Title
        title_label = tk.Label(self.ns_window, text='Create Store')
        title_label.grid(column=1, row=0, columnspan=5, pady=15)
        
        # Store input
        store_name_label = tk.Label(self.ns_window, text='Store name : ')
        store_name_label.grid(row=2, column=1, padx=10)
        self.store_name_box = tk.Entry(self.ns_window, width=30)
        self.store_name_box.grid(column=2, row=2, columnspan=4, pady=5, sticky='W')
        if update:
            self.store_name_box.insert(0, update['name'])
        
        # Branch / id input
        num_label = tk.Label(self.ns_window, text='Branch / ID number: ')
        num_label.grid(row=3, column=1, padx=10)
        self.num_box = tk.Entry(self.ns_window, width=30)
        self.num_box.grid(column=2, row=3, columnspan=4, pady=5, sticky='W')
        if update:
            self.branch_box.insert(0, update['num'])

        # Contact input
        contact_label = tk.Label(self.ns_window, text='Contact name : ')
        contact_label.grid(row=4, column=1, padx=10)
        self.contact_box = tk.Entry(self.ns_window, width=30)
        self.contact_box.grid(column=2, row=4, columnspan=4, pady=5, sticky='W')
        if update:
            self.contact_box.insert(0, update['contact'])

        # Phone input
        phone_label = tk.Label(self.ns_window, text='Contact phone : ')
        phone_label.grid(row=5, column=1, padx=10)
        self.phone_box = tk.Entry(self.ns_window, width=30)
        self.phone_box.grid(column=2, row=5, columnspan=4, pady=5, sticky='W')
        if update:
            self.phone_box.insert(0, update['phone'])
        
        # Email input
        email_label = tk.Label(self.ns_window, text='Contact email : ')
        email_label.grid(row=6, column=1, padx=10)
        self.email_box = tk.Entry(self.ns_window, width=30)
        self.email_box.grid(column=2, row=6, columnspan=4, pady=5, sticky='W')
        if update:
            self.email_box.insert(0, update['email'])

        # Pass input
        pw_label = tk.Label(self.ns_window, text='Password : ')
        pw_label.grid(row=7, column=1, padx=10)
        self.pw_box = tk.Entry(self.ns_window, show='●', width=30)
        self.pw_box.grid(column=2, row=7, columnspan=4, pady=5, sticky='W')
        #if not update:
            #self.pw_box.insert(0,'Minimum 8 characters, at least one number, one uppercase and one symbol')
        if update:
            self.pw_box.insert(0, update['pass'])

        # Confirm Pass input
        confirm_pw_label = tk.Label(self.ns_window, text='Confirm password : ')
        confirm_pw_label.grid(row=8, column=1, padx=10)
        self.confirm_pw_box = tk.Entry(self.ns_window, show='●', width=30)
        self.confirm_pw_box.grid(column=2, row=8, columnspan=4, pady=5, sticky='W')
        if update:
            self.pw_box.insert(0, update['pass'])

        # Labor
        labor_label = tk.Label(self.ns_window, text='Avg labor : ')
        labor_label.grid(row=9, column=1)
        self.labor_box = tk.Entry(self.ns_window, width=6)
        self.labor_box.grid(column=2, row=9, sticky='W')
        if update:
            self.labor_box.insert(0, update['labor'])

        # Markup
        markup_label = tk.Label(self.ns_window, text='Markup % : ')
        markup_label.grid(row=9, column=2, columnspan=2)
        self.markup_box = tk.Entry(self.ns_window, width=6)
        self.markup_box.grid(column=3, row=9, sticky='E')
        if update:
            self.markup_box.insert(0, update['markup'])

        # Buttons
        submit_button = tk.Button(self.ns_window, text='Submit', command=self.__submit_new_store__, width=10)
        submit_button.grid(column=2, row=10, pady=15, padx=5, sticky='E')
        cancel_button = tk.Button(self.ns_window, text='Back', command=self.__back_button__, width=10)
        cancel_button.grid(column=3, row=10, pady=15, padx=5, sticky='W')
        
        self.ns_window.mainloop()

    def __submit_new_store__(self):
        store = {}
        stores = __load_json__('stores.json')
        
        # Error checking
        # TODO put this in the boxes, but make it disappear when box selected and red font
        error_message = []
        name = self.store_name_box.get()
        if not name:
            error_message.append('Missing store name. ')
        num =  self.num_box.get()
        branch = str(name + num).replace(' ', '') # remove spaces and combine name and num
        for sid in stores:
            if stores[sid]['branch'] == branch:
                messagebox.showinfo('Duplicate Store', branch + ' already exists in database. Please choose a different store name / id or click \'Back\' to return to login screen.')
                return
        contact = self.contact_box.get()
        if not contact:
            error_message.append('Missing contact name. ')  
        phone = self.__parse_phone__(self.phone_box.get())
        if not phone:
            error_message.append('Please enter a 10-digit phone number. ')
        email = self.__parse_email__(self.email_box.get())
        if not email:
            error_message.append('Please enter a a valid email address. ')
        pw = self.__vette_pw__(self.pw_box.get())
        if isinstance(pw, list):
            error_message.append('Password does not meet requirements : ')
            for line in pw:
                error_message.append('   - ' + line)
        if not pw == self.confirm_pw_box.get():
            error_message.append('Passwords do not match. ')  
        try:
            labor = float(self.labor_box.get())
        except ValueError:
            error_message.append('Avg. labor must be a number.')
        try:
            markup = float(self.markup_box.get())
        except ValueError:
            error_message.append('Markup percentage must be a number.')
        if error_message:
            message = ''
            for line in error_message:
                message += line + '\n'
            messagebox.showinfo('Missing or incorrect information', message)
            return
        
        # TODO 'update' implementation
        store['name'] = name
        store['num'] = num
        store['branch'] = branch
        store['contact'] = contact
        store['email'] =  email
        store['pass'] = pw
        store['phone'] = phone 
        store['db_file'] = self.__get_db_max__()
        store['id'] = '1' + store['db_file'].split('db')[1][:-5]
        store['encription'] = None # TODO
        stores[store['id']] = store # add the new store to the database
        self.__update_json__(stores, 'stores.json', 'item')
        store['labor'] = labor # labor and markup are not added to the stores db, only the db file
        store['markup'] = markup
        self.store = Store(self.__create_database__(store))
        messagebox.showinfo('Create Store','Store ' + store['branch'] + ' sucessfully created!')
        self.ns_window.destroy()
        return

    def __back_button__(self):
        self.ns_window.destroy()
        self._login()

    def __create_database__(self, store):
        """Creates a new Store database"""
        db = {  'creds':       {'db_file':  store['db_file'], 
                                'id':       store['id'], 
                                'branch':   store['branch']}, 
                'costs':       {'labor' :   store['labor'], 
                                'markup' :  store['markup']}, 
                'sql_creds' :  {'host' :    None, 
                                'database' : store['branch'], 
                                'user' :    None, 
                                'pass' :    None}, 
                'ingredients': {}, 
                'recipes' :    {}}
        self.__update_json__(db, store['db_file'])
        return db        

    def __get_db_max__(self):
        """Searches the current directory for files starting with 'db'
        Returns the next iteration after highest db number"""
        filenames = os.listdir()
        filenames.sort()
        maxnum = 0
        for name in filenames:
            if name.startswith('db'):
                maxnum = int(name.split('db')[1][:-5])     
        if maxnum < 10:
            outfile = 'db00'
        elif maxnum < 100:
            outfile = 'db0'
        else:
            outfile = 'db'
        return outfile + str(maxnum + 1) + '.json'
 
    def __parse_phone__(self, p):
        """Ensure phone number is valid"""
        phone_re = re.compile(r'\d{10}$')
        seq_type= type(p)
        p = seq_type().join(filter(seq_type.isdigit, p))
        if phone_re.match(p):
            return p
        return None    
            
    def __parse_email__(self, e):
        """Ensure email address is formatted correctly"""
        email_re = re.compile(r'''(
        [a-zA-Z0-9._%+-]+               # username
        @                               # @
        [a-zA-Z0-9.-]+                  # domain name
        (\.[a-zA-Z]{2,4})               # dot something
        )''', re.VERBOSE)
        if email_re.match(e):
            return e
        return None

    def __vette_pw__(self, p):
        """Enssure password meets requirements"""
        length_error = len(p) < 8
        digit_error = re.search(r"\d", p) is None
        uppercase_error = re.search(r"[A-Z]", p) is None
        lowercase_error = re.search(r"[a-z]", p) is None
        symbol_error = re.search(r"[ !#$%&'()*+,-./[\\\]^_`{|}~"+r'"]', p) is None
        password_ok = not ( length_error or digit_error or uppercase_error or lowercase_error or symbol_error )
        errors = {'password_ok' : password_ok,
            'length less than 8 characters' : length_error,
            'missing digit' : digit_error,
            'missing uppercase' : uppercase_error,
            'missing lowercase' : lowercase_error,
            'missing symbol' : symbol_error,}
        if password_ok:
            return p
        else:
            error_list = []
            for k, v in errors.items():
                if v:
                    error_list.append(k)
            return error_list

    def __choices__(self, title, choices):      
        """
        Creates and validates choices in a numbered list\n
        returns int value or none if last option is choosen\n 
        variable 'choices' is a list
        """
        print(title)
        for num, choice in enumerate(choices):
            print(f'{num + 1}. {choice}')
        print(f'{len(choices) + 1}. Back')
        while True:
            try:
                ch = int(input('Choose an option. '))
                if ch > len(choices) + 1:
                    raise ValueError
                elif ch == len(choices) + 1:
                    return None
                else:
                    return ch
            except ValueError:
                print('Invalid entry')

    def __pause__(self):
        pause = input('\n< Press \'enter\' to go back > ')


    def __update_store__(self, updated_store):
        """Updates data in the current Store. Database will not be updated until __save__ is called"""
        global LABOR
        global MARKUP
        db = {  'creds':       {'db_file': updated_store['db_file'], 
                                'id': updated_store['id'], 
                                'branch': updated_store['branch']}, 
                'costs':       {'labor' : LABOR, 
                                'markup' :MARKUP}, 
                'sql_creds' :  self.store.sql_creds, 
                'ingredients': __encode_json__(self.store.ingredients), 
                'recipes' :    __encode_json__(self.store.recipes)}
        self.__update_json__(db, updated_store['db_file'])
        return db

    def __update_json__(self, dct, filepath, dict_name=None):
        if not dict_name:
            dict_name = 'item'
        with open(filepath, 'w') as f:
            json.dump(dct, f)
        print('\nDatabase updated.')
        #print(f'Updated {len(dct)} {dict_name}(s) in {filepath}')
        time.sleep(1)

    def __print_frame__(self, data, keys, title=None):
        """data must be a dict of name and objects that contain keys"""
        out = {}
        for key in keys:
            lst = []
            for dat in data:
                lst.append(data[dat].item_dict[key])
            out[key] = lst
        pd.set_option('display.max_rows', None)  
        df = pd.DataFrame.from_dict(out)
        if title:
            # TODO This doesn't work
            df.style.set_caption(title)
        print(df)

    def __get_data_ranges__(self, req_keys, optional_keys=None):
        """Return user-inputted ranges needed to parse excel data"""
        while True:
            try:
                row_range = str(input('Row range <eg. 1-10> '))
                row_min = row_range.split('-')[0]
                row_max = row_range.split('-')[1]
                row_range = (int(row_min), int(row_max))
                break
            except:
                print(f'Error, row range must be formatted <min-max>')  
        col_list = []
        col_dict = {}
        for item in req_keys:
            while True:
                try:
                    main_col = int(input(f'Column containing {item}? '))
                    col_dict[item] = main_col
                    col_list.append(main_col)
                    break
                except ValueError:
                    print('Error, column numbers must be integers.')
        opt_dict = {}
        if optional_keys:
            print('\n---Optional data---\nLeave blank if not applicable')    
            for opt in optional_keys:
                while True:
                    try:
                        opt_col = input(f'Column containing {opt}? ')
                        if opt_col:
                            opt_dict[opt] = int(opt_col)
                            col_list.append(int(opt_col))
                            break
                    except ValueError:
                        print('Error, column must be an integer.')
                    break
            
        col_range = (min(col_list), max(col_list))
        return row_range, col_range, col_dict, opt_dict

    def __check_ingredient__(self, ingredient, recipe):
        """Check if ingredient already exists in a recipe"""
        if ingredient in self.store.ingredients:
            return self.store.ingredients
        else:
            print(f'\nNEW INGREDIENT \'{ingredient}\' found in recipe \'{recipe.title()}\'.')
            return self.add_ingredients(ingredient)

    def __save__(self):
        self.__update_json__(self.store._json(), self.store.db_file)

    # Terminal interface
    def login(self, args=None):
        """Login section"""
        # TODO encript stores.json
        stores = __load_json__('stores.json')
        while True:
            os.system('cls')
            login = 'Login'
            if args:
                login = login + ' (Credentials provided in command line)'
            ch = self.__choices__('==== FOOD COSITNG ====\n',['New store', login])
            if not ch:
                os.system('cls')
                sys.exit()
            elif ch == 1:
                store = self.create_store(stores)
                self.__pause__()
                if not store:
                    continue
                else:
                    return store
            elif ch ==2:
                if args:
                    email = args[1]
                    pw = args[2]
                else:
                    os.system('cls')
                    print('==== LOGIN ====\n')
                    email = input('email ')
                    pw = input('password ')
            for store in stores:
                if stores[store]['email'] == email and stores[store]['pass'] == pw:
                    print('\nLogin Success.')
                    time.sleep(1)
                    return __load_json__(stores[store]['db_file'])
            print('\nIncorrect email and/or password.')
            args = None
            time.sleep(2)

    # Terminal interface
    def create_store(self, stores, update=None):
        """Create a new Store or update information on an existing Store"""
        """Optional 'update' set to store_id"""
        if not update:
            store = {}
        keys = [('Store name', 'name'), 
                ('Branch number / Store ID', 'num'),
                ('Contact full name' ,'contact'),
                ('Contact phone', 'phone'), 
                ('Contact email', 'email'),
                ('Password (Minimum 8 characters, at least one number, one uppercase and one symbol)', 'pass')]
        
        # Menu title
        os.system('cls')
        rv = ''
        title = 'NEW STORE'
        if update:
            current_store = stores[update]
            #print(f'{current_store=}')
            title = 'UPDATE ' + current_store['branch']
        print(f'==== {title} ====\n\nStore Information:\n(type \'done\' at any time to go back)\n')
        if update:
            print('(leave selections blank to retain values)')
        
        # Store information input
        for key in keys:
            val = None
            if update:
                retain_value = current_store[key[1]]
                rv = (f' (current value : {retain_value})')
            while not val:
                val = input(f'{key[0]}{rv}: ')
                if not val and update:
                    val = retain_value
                    print('  value retained')
                    break
                if val == 'done':
                    if update:
                        return stores
                    return None
                if key[1] == 'phone':
                    val = self.__parse_phone__(val)
                    if not val:
                        print('Please enter a 10-digit phone number')
                        continue
                elif key[1] == 'email':
                    val = self.__parse_email__(val)
                    if not val:
                        print('Please enter a valid email')
                        continue
                elif key[1] == 'pass':
                    val = self.__vette_pw__(val)
                    if isinstance(val, list):
                        print(f'Password missing requirement(s): \n{val}')
                        val = None
                        continue
            if key[1] == 'pass' and not update:
                check = None
                while not check:
                    check = input('Confirm password: ')
                    if not check:
                        return None
                    if check != val:
                        print('Passwords do not match. (leave blank to go back)')
                        check = None
            if update:
                stores[update][key[1]] = val
                if key[1] == 'num':
                    branch = current_store['name'] + val
                    if branch not in stores:
                        stores[update]['branch'] = branch
                    else:
                        print('A store with this name already exists, please use the login option or try again')
                        return None
            else:
                store[key[1]] = val
                if key[1] == 'num':
                    branch = store['name'] + val
                    if branch not in stores:
                        store['branch'] = branch
                    else:
                        print('A store with this name already exists, please use the login option or try again')
                        return None
        
        # Confirm store creation
        if not update:
            print('\nSTORE : ' + store['branch'])
            print('Contact : ' + store['contact'])
            print('Phone : ' + store['phone'])
            print('Email : ' + store['email'])
            yn = input('\nCreate store? (Y/N)')
            if yn.lower() == 'y':
                store['db_file'] = self.__get_db_max__()
                store['id'] = '1' + store['db_file'].split('db')[1][:-5]
                store['encription'] = None # TODO
                stores[store['id']] = store # add the new store to the database
                self.__update_json__(stores, 'stores.json', 'item')
                print('Store ' + store['branch'] + ' sucessfully created!')
                time.sleep(2)
                return self.create_database(store)
            else:
                print('Store creation cancelled.')
                time.sleep(2)
                return None
        else:
            return stores
        
    # Terminal interface
    def create_database(self, store):
        """Creates a new Store database"""
        store_id = store['branch']
        os.system('cls')
        print(f'==== STORE ID : {store_id} ====\n\nInitial values:\n')
        while True:
            try:
                labor = int(input('Average labor cost? '))
                break
            except ValueError:
                print('Labor cost must be an integer.')
        while True:
            try:
                markup = float(input('Recipe markup percentage? '))
                break
            except ValueError:
                print('Labor cost must be a number.')
        db = {  'creds':       {'db_file': store['db_file'], 
                                'id': store['id'], 
                                'branch': store['branch']}, 
                'costs':       {'labor' : store['labor'], 
                                'markup' : store['markup']}, 
                'sql_creds' :  {'host' : None, 
                                'database' : store['branch'], 
                                'user' : None, 
                                'pass' : None}, 
                'ingredients': {}, 
                'recipes' :    {}}
        self.__update_json__(db, store['db_file'])
        return db

    def main_menu(self):
        """Main Menu"""
        choice = None
        while True:
            os.system('cls')
            print(f'===== STORE ID: {self.store.branch} =====\n')
            print('1. Update store information')
            print('2. Add / Update / Remove item')
            print('3. Update costs')
            print('4. Search database')
            print('5. List all')
            print('6. Connect to database')
            print('7. Save')
            print('8. Quit')
            while True:
                try:
                    choice = int(input('Choose an option. '))
                    if choice == 1:
                        os.system('cls')
                        ch = self.__choices__(f'===== STORE ID: {self.store.branch} =====\n', ['Update information', 'Remove Store'])
                        if not ch:
                            break
                        stores =  __load_json__('stores.json')
                        if ch == 1:
                            updated_stores = self.create_store(stores, update=self.store.id)
                            self.__update_json__(updated_stores, 'stores.json')
                            self.store = Store(self.__update_store__(updated_stores[self.store.id]))
                            print('\nStore information updated.')
                            time.sleep(2)
                            break
                        elif ch == 2:
                            yn = input(f'Removing store \'{self.store.branch}\' cannont be undone, are you sure? (Y/N) ')
                            if yn.lower() == 'y':
                                stores.pop(self.store.branch)
                                self.__update_json__(stores, 'stores.json')
                                os.remove(self.store.db_file)
                                print(f'Store removed. Returning to log in screen.')
                                time.sleep(2)
                                self.__init__()
                            else:
                                break
                    elif choice == 2:
                        os.system('cls')
                        ch = self.__choices__('~~~~ ADD / UPDATE / REMOVE ITEM ~~~~\n', ['Add / Update', 'Remove'])
                        if not ch:
                            break
                        elif ch == 1:
                            os.system('cls')
                            cho = self.__choices__('~~~ ADD / UPDATE ITEM ~~~\n', ['Ingredient', 'Recipe'])
                            if not cho:
                                break
                            elif cho == 1:
                                self.add_ingredient()
                                self.__pause__()
                                break
                            elif cho == 2:
                                self.add_recipe()
                                self.__pause__()
                                break
                            self.__pause__()
                            break
                        elif ch == 2:
                            # Remove item
                            search = self.search_item(self.store.all_items)
                            if not search:
                                break
                            class_name = search.__class__.__name__
                            name = search.name
                            if class_name == 'Ingredient':
                                keys = self.xl['ing_req_cats'] + self.xl['ing_opt_cats']
                                self.__print_frame__({search.name : search}, keys)
                            elif class_name == 'Recipe':
                                search.print_recipe()
                            cho = input('\nAre you sure you want to remove item? (Y/N) ')
                            if cho.lower() == 'y':
                                print(f'\n{name.capitalize()} removed from database.')
                                if class_name == 'Ingredient':
                                    self.store.ingredients.pop(name)
                                elif class_name == 'Recipe':
                                    self.store.recipes.pop(name)
                                time.sleep(2)
                                break
                            self.__pause__()
                            break          
                    elif choice == 3:
                        os.system('cls')
                        print('<<< COST VARIABLES >>>\n\n(leave blank to retain current value)')
                        global LABOR 
                        global MARKUP 
                        LABOR = self.update_cost('labor', LABOR)
                        MARKUP = self.update_cost('markup', MARKUP)
                        print('\nCost values updated.')
                        time.sleep(2)
                        break
                    elif choice == 4:
                        os.system('cls')
                        # Search database
                        while True:
                            result = self.search_item(self.store.all_items)
                            class_name = result.__class__.__name__
                            if not result:    
                                break
                            elif class_name == 'Ingredient':
                                print()
                                self.__print_frame__({result.name : result}, self.xl['ing_req_cats'] + self.xl['ing_opt_cats'])
                                self.__pause__()
                            elif class_name == 'Recipe':
                                print()
                                result.print_recipe()
                                self.__pause__()
                            else:
                                continue
                        break
                    elif choice == 5:
                        os.system('cls')
                        # List all items in database
                        while True:
                            os.system('cls')
                            ch = self.__choices__('--- LIST ALL ---\n', ['Ingredients', 'Recipes'])
                            if not ch:
                                break
                            elif ch == 1:
                                if not self.store.ingredients:
                                    print('\nNo ingredients loaded.')
                                    self.__pause__()
                                    break
                                print('\n=== INGREDIENTS ===\n')
                                # TODO get title of table to work
                                self.__print_frame__(self.store.ingredients, self.xl['ing_req_cats'], '=== INGREDIENTS ===')
                                print(f'\nTotal {len(self.store.ingredients)} ingredient(s)')
                                self.__pause__()
                                break
                            elif ch == 2:
                                if not self.store.recipes:
                                    print('\nNo recipes loaded.')
                                    self.__pause__()
                                    break
                                print('\n*** RECIPES ***')
                                for recipe in self.store.recipes:
                                    self.store.recipes[recipe].print_recipe()
                                self.__pause__()
                                break
                            break
                        break
                    elif choice == 6:
                        os.system('cls')
                        # Import data from excel or connect to an external database
                        ch = self.__choices__('___ CONNECT TO DATABASE ___\n', ['Import Excel', 'Connect to JSON database', 'Connect to SQL database'])
                        if not ch:
                            break
                        elif ch == 1:
                            self.connect_xl()
                            break
                        elif ch == 2:
                            ingredients, recipes = connect_sql(sql_creds)
                        if not ingredients:
                            print('No ingredients loaded from database.')
                        if not recipes:
                            print('No recipes loaded from database.')
                        pause()
                        break
                    elif choice == 7:
                        # Save
                        self.__save__()
                        break
                    elif choice == 8:
                        # Quit
                        yn = input('\nAny changes made since your last save will be lost, are you sure? (Y/N) ')
                        if yn.lower() == 'y':
                            os.system('cls')
                            sys.exit()
                        else:
                            self.main_menu()
                    else:
                        raise ValueError
                except ValueError:
                    print('Invalid input.')                           

    def add_ingredient(self, name=None, uuid=None):
        """Add a new ingredient to database\n
        If ingredient exists, user is prompted to update ingredient information"""
        ret = False # ret is a bool that is used to return to add_recipe if a name is passed to add_ingredient
        while True:
            print('\n+++ ADD / UPDATE INGREDIENT +++\n(leave blank to return to main menu)\n(type \'list\' to list all ingredients)\n\n(enter an existing ingredient to update)')
            if name:
                ret = True
                print(f'Ingredient : \'{name}\'')
            if not name:
                name = input('Name of ingredient? ')
                name = name.lower() # ingredients must have lower case names
            if name == 'list':
                self.__print_frame__(self.store.ingredients, self.xl['ing_req_cats'], '=== INGREDIENTS ===')
                print(f'\nTotal {len(self.store.ingredients)} ingredient(s)')
                name = None
                continue
            elif not name:
                return

            # Ingredient already exists in database
            elif name in self.store.ingredients.keys():
                choice = self.__choices__(f'\n\'{name}\' already exists in the database.\n+++ UPDATE INGREDIENT +++\n', ['Update'])
                while True:
                    if not choice:
                        return
                    elif choice == 1:
                        self.store.ingredients[name] = self.update_ingredient(self.store.ingredients[name])
                        self.store.ingredients[name]._print_ingredient()
                        return
            
            # Enter new ingredient information
            unit = None
            quantity = None
            price = None
            while not unit:
                unit = input('What unit of measurement? ')
            while not quantity:
                try:
                    quantity = float(input(f'How many total {unit}? '))
                except ValueError:
                    print('Invalid quantity, please choose a number.')
            while not price:
                try:
                    price = float(input('Price? '))
                except ValueError:
                    print('Invalid price, please choose a number.')
            self.store.ingredients[name] = Ingredient(name, unit, quantity, price, uuid=uuid)
            print(f'\'{name.capitalize()}\' added to ingredients database.\n')
            self.store.ingredients[name]._print_ingredient()
            name = None
            if ret:
                return

    def update_ingredient(self, ingredient):
        """Update an existing ingredient"""
        # TODO DRY, iterate through .factors (i.e. .name, .unit, .price) pass each to a function?
        # TODO this does not iterate through optional information
        print(f'\n+++ UPDATE INGREDIENT \'{ingredient.name.upper()}\' +++\n(Leave choices blank if you wish to retain original values)')    
        while True:
            unit = input(f'Current unit of measurement : \'{ingredient.unit}\', new unit? ')
            if not unit:
                unit = ingredient.unit
                print('    Unit retained.')
                break
            else:
                try:
                    unit = float(unit)
                    print('    Unit updated.')
                    break
                except ValueError:
                    print('Invalid entry, please enter a number.')
        while True:
            quantity = input(f'Current total quantity : {ingredient.quantity} {ingredient.unit}, how many {unit}? ')
            if not quantity:
                quantity = ingredient.quantity
                print('    Quantity retained.')
                break
            else:
                try:
                    quantity = float(quantity)
                    print('    Quantity updated.')
                    break
                except:
                    print('Invalid entry, please enter a number.')
        while True:
            price = input(f'Current price : {ingredient.price}, new price? ')
            if not price:
                price = ingredient.price
                print('    Price retained.')
                break
            else:
                try:
                    price = float(price)
                    print('    Price updated.')
                    break
                except:
                    print('Invalid entry, please enter a number.')
        print(f'\'{ingredient.name}\' updated.\n')
        return Ingredient(ingredient.name, unit, quantity, price)

    def add_recipe(self):
        """Add a a new Recipe"""
        while True:
            print('\n*** ADD / UPDATE RECIPE ***\n(Leave blank to return to main menu)\n(enter an existing recipe to update) ')
            name = input('Name of recipe? ')
            name = capwords(name) # Recipes must have capitalized names
            if not name:
                return
            # Check if recipe already exists and update
            if name in self.store.recipes:
                while True:
                    print(f'\n\'{name}\' already exists in the database.')
                    choice = self.__choices__('\n\n+++ UPDATE RECIPE +++', ['Update'])
                    if not choice:
                        return 
                    if choice == 1:
                        self.store.recipes[name] = self.update_recipe(recipes[name])
                        self.store.recipes[name].print_recipe()                    
                        return
            
            print(f'\n+++ NEW RECIPE \'{name}\' +++')
            # Is this recipe used as an ingredient in other recipes?
            is_ing = False
            yn = input('Is this recipe used as an ingredient in other recipes? (Y/N) ')
            if yn == 'y':
                is_ing = True
            
            # How long to make? (preptime)
            while True:
                try:
                    preptime = float(input('Prep time? (minutes) '))
                    break
                except ValueError:
                    print('Invalid prep time. Input must be a number.')

            # Add ingredients
            inglist = []    
            ing = ''
            while ing != 'done':
                ing = ''
                while True:
                    ing = input('Ingredient? (type \'done\' when finished) ')
                    ing = ing.lower()
                    if not ing:
                        continue
                    if ing == 'done':
                        self.store.recipes[name] = Recipe(name, inglist, preptime, is_ingredient=is_ing)
                        print(f'\n\'{name}\' added to recipe database.')
                        self.store.recipes[name].print_recipe()
                        if self.store.recipes[name].is_ingredient:
                            self.store.ingredients[name.lower()] = self.store.recipes[name].as_ingredient
                            print(f'\n\'{name.lower()}\' added to ingredient database.\n')
                            self.store.ingredients[name.lower()]._print_ingredient()
                        return
                    elif ing in self.store.ingredients: 
                        print('    Ingredient found')
                        unit = self.store.ingredients[ing].unit
                        while True:
                            try:
                                quantity = float(input('Quantity of ' + unit + '? '))
                                break    
                            except ValueError:
                                print('Invalid quantity. Input must be a number.')
                        
                        print(f'\'{ing}\' added to recipe \'{name}\'')
                        inglist.append((self.store.ingredients[ing], quantity))
                    else:
                        while True:
                            yn = input('Ingredient not found. Would you like to create this ingredient? (Y/N) ')
                            if yn.lower() == 'n':
                                break
                            else:
                                self.add_ingredient(ing)
                                print(f'\n*** RECIPE {name.upper()} ***\n\'{ing}\' added to \'{name}\'')
                                unit = self.store.ingredients[ing].unit
                                while True:
                                    try:
                                        quantity = float(input('Quantity of ' + unit + '? '))
                                        break    
                                    except ValueError:
                                        print('Invalid quantity. Input must be a number.')    
                                inglist.append((self.store.ingredients[ing], quantity))
                                break

    def update_recipe(self, recipe):
        """Update an existing Recipe"""
        print(f'\n*** UPDATE RECIPE \'{recipe.name.upper()}\' ***\n(Leave choices blank if you wish to retain original values\n(Type \'remove\' to remove an ingredient)')
        preptime = input(f'Current prep time : {recipe.preptime} minutes, new prep time? ')
        if not preptime:
            print('    Prep time retained.')
            preptime = recipe.preptime
        pop_list = []
        for index, (ingredient, quantity) in enumerate(recipe.ingredients):
            quan = input(f'Ingredient \'{ingredient.name}\', {quantity} {ingredient.unit}. New quantity? ')
            if quan == 'remove':
                print(f'    \'{ingredient.name}\' has been removed.')
                pop_list.append(index)
            elif not quan:
                print('    Quantity retained.')
                recipe.ingredients[index] = (ingredient, quantity)
            else:
                # TODO don't like this
                quanchk = False
                while quanchk is False:
                    try:
                        quantity = float(quan)
                        quanchk = True
                        recipe.ingredients[index] = (ingredient, quantity)
                        print('    Quantity updated.')
                    except:
                        quan = input('Invalid quantity, please enter a number. ')
        # TODO This section is redundant from get_recipe, break into its own function?
        inglist = []
        ing = ''
        while True:
            ing = ''
            while True:
                ing = input('Ingredient? (type \'done\' to to complete recipe) ')
                
                if ing.lower() == 'done':
                    print(f'\'{recipe.name}\' updated.')
                    for pop in pop_list:
                        recipe.ingredients.pop(pop) #  "- Magnitude
                    return Recipe(recipe.name, recipe.ingredients, preptime)

                
                elif ing in ingredients: 
                    print('    Ingredient found')
                    unit = ingredients[ing].unit
                    while True:
                        try:
                            quantity = float(input('Quantity of ' + unit + '? '))
                            break    
                        except ValueError:
                            print('Invalid quantity. Input must be a number.')
                    
                    print(f'\'{ing}\' added to recipe \'{recipe.name}\'')
                    recipe.ingredients.append((ingredients[ing], quantity))
                else:
                    while True:
                        yn = input('Ingredient not found. Would you like to create this ingredient? (Y/N) ')
                        if yn.lower() == 'n':
                            break
                        else:
                            self.store.ingredients = self.add_ingredients(ing)
                            print(f'\n*** RECIPE {recipe.name.upper()} ***\n\'{ing}\' added to \'{recipe.name}\'')
                            unit = self.store.ingredients[ing].unit
                            while True:
                                try:
                                    quantity = float(input('Quantity of ' + unit + '? '))
                                    break    
                                except ValueError:
                                    print('Invalid quantity. Input must be a number.')    
                            recipe.ingredients.append((self.store.ingredients[ing], quantity))
                            break

    def update_cost(self, term, var):
        """Used for updating global LABOR and MARKUP variables to new values"""
        while True:
            out = input(f'Current {term} : {var} - new {term}? ')
            if not out:
                print(f'    {term.capitalize()} retained.')
                return var
            try:
                return float(out)
            except ValueError:
                print('Unrecognized input. Please enter a number.')

    def search_item(self, search_dict):
        """Search the database for an item"""
        os.system('cls')
        print(f'<<< SEARCH DATABASE >>>\n\n(Leave blank to return to main menu)')
        while True:
            search = input(f'Search for? ')
            if not search:
                return None
            result_dict = {}
            num = 1
            for item in search_dict:
                name = item.lower()
                if search.lower() in name:
                    result_dict[num] = (item, search_dict[item])
                    num = num + 1
            if not result_dict:
                print('No results found.')
                continue
            print('\n<<< SEARCH RESULTS >>>')
            for result in result_dict:  
                print(f'{result}. {result_dict[result][0]}')
            while True:
                choice = input(f'Choose a number from the above list (Leave blank to return to main menu) ')
                if not choice:
                    return None
                try:
                    choice = int(choice)
                    if choice in range(1, len(result_dict) + 1):
                        return result_dict[choice][1]
                    else:    
                        raise ValueError
                except ValueError:
                    print('Invalid choice.')
                    continue
            print(f'No items found.')

    def connect_xl(self):
        """Connect to an external excel sheet"""
        # Load workbook
        while True:
            types = ['xls', 'xlsx', 'xlsm']
            book = input('\nExcel path and filename (leave blank to exit) : ')
            if not book:
                return
            if not book[-6:].split('.')[1] in types:
                print('Not a vaild Excel file.')
                continue
            elif not os.path.exists(book):
                print('File does not exist.')
                continue
            try:
                wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
            except Exception as e:
                print(f'An error has occured. {e}')
            break
        # TODO DRY    
        # Load sheet
        while True:
            sheet = input('Ingredient sheet name (leave blank to skip, type \'list\' to list sheets.) ')
            if not sheet:
                break
            elif sheet.lower() == 'list':
                for ws in wb.sheetnames:
                    print(ws)
                continue
            try:
                sheet = wb[sheet]
                break
            except KeyError:
                print(f'Sheet \'{sheet}\' does not exist.')        
        while True:
            if not sheet:
                break
            try:
                self.find_xl_ingredients(sheet)
                break
            except Exception as e:
                print(e)
                break
        while True:
            # Recipe input
            sheet = input('Recipie sheet name (Leave blank to exit, type \'list\' to list sheets.) ')
            if not sheet:
                return
            elif sheet.lower() == 'list':
                for ws in wb.sheetnames:
                    print(ws)
                continue
            try:
                sheet = wb[sheet]
                self.find_xl_recipes(sheet)
                return
            except KeyError:
                print(f'Sheet \'{sheet}\' does not exist.')

    def find_xl_ingredients(self, ws):
        """Load ingredients from an excel sheet"""
        """
        Required columns:   name    unit    quanty  price
        i.e:                cheese  slices  20      3.99
                            mustard oz      24      4.99
                            bun     buns    8       2.50
        """
        rows, cols, reqs, optional = self.__get_data_ranges__(self.xl['ing_req_cats'], self.xl['ing_opt_cats'])
        count = 0
        err_rows = []
        # TODO if ingredient already exists and has new data? Might not be necessary
        for row in range(rows[0], rows[1] + 1):
            name = None
            unit = None
            quantity = None
            price = None
            vendor = None
            notes = None
            cal = 0
            serv = 1
            for col in range(cols[0], cols[1] + 1):
                val = ws.cell(row=row, column=col).value
                # TODO DRY Dont like hard-coded names here, use indicies?
                if col == reqs['name']:
                    name = val.lower()
                elif col == reqs['unit']:
                    unit = val
                elif col == reqs['quantity']:
                    quantity = val
                elif col == reqs['price']:
                    price = val
                elif optional['vendor'] and col == optional['vendor']:
                    if val:
                        vendor = val.lower()
                    else:
                        vendor = None
                elif optional['notes'] and col == optional['notes']:
                    notes = val
                elif optional['calories'] and col == optional['calories']:
                    cal = val
                elif optional['servings'] and col == optional['servings']:
                    serv = val
            try:
                self.store.ingredients[name] = Ingredient(name, unit, quantity, price, vendor=vendor, notes=notes, calories=cal, servings=serv)
                count = count + 1
            except Exception as e:
                err_rows.append(row)
                continue
        print(f'\nAdded {count} ingredient(s). ', end='')
        if len(err_rows) > 0:
            print(f'Ingredients in row(s) {err_rows} could not be updated due to missing data requirements.')
        else:
            print('')

    def find_xl_recipes(self, ws):
        """Load recipes from an excel sheet\n
        Required columns with example:   
        name    ingredients quantity    preptime\n
        Burger  patty       1           5\n
        Burger  cheese      2           5\n
        Burger  bun         2           5\n
        Hotdog  frank       1           2\n
        Hotdog  ketchup     4           2\n
        Hotdog  bun         1           2\n
        """
        rows, cols, reqs, optional = self.__get_data_ranges__(self.xl['rec_req_cats'], self.xl['rec_opt_cats'])
        count = 0
        err_rows = []
        recipe_dict = {}
        ing_list_tuple = []
        name = None
        ingredient = None
        quantity = None
        prep = None
        notes = None
        yld = None
        for row in range(rows[0], rows[1] + 2): # Needs to read one extra row to add preptime
            for col in range(cols[0], cols[1] + 1): 
                val = ws.cell(row=row, column=col).value
                # TODO DRY
                if row == rows[1] +1 and col == cols[1]:
                    ing_list_tuple.append(prep)
                    if notes:
                        ing_list_tuple.append(notes)
                    elif yld:
                        ing_list_tuple.append(yld)        
                    self.store.recipes[name] = ing_list_tuple
                    #print(f'RECIPE {name} COMPLETE')

                elif col == reqs['name']:
                    if val not in self.store.recipes:
                        if name:
                            # Recipe complete
                            ing_list_tuple.append(prep)
                            if notes:
                                ing_list_tuple.append(notes)
                            elif yld:
                                ing_list_tuple.append(yld)
                            
                            self.store.recipes[name] = ing_list_tuple
                            #print(f'RECIPE {name} COMPLETE')
                            prep = None
                            notes = None
                            yld = None
                        name = val
                        if val is None:
                            continue
                        print('\nNEW RECIPE ' + name)
                        self.store.recipes[name] = []
                        ing_list_tuple = []
                elif col == reqs['ingredients']:
                    try:
                        ingredient = val.lower()
                    except:
                        pass
                elif col == reqs['quantites']:
                    quantity = val
                elif col == reqs['preptime']:
                    prep = ('preptime', val)
                elif col == optional['notes']:
                    if val:
                        notes = ('notes', val)
                elif col == optional['yield']:
                    if val:
                        yld = ('yld', val)
                if ingredient and quantity:
                    ing_list_tuple.append((ingredient, quantity))
                    ingredient = None
                    quantity = None
        for recipe in self.store.recipes:
            name = None
            notes = None
            yld = None
            ings = self.store.recipes[recipe]
            ing_list = []
            for i in ings:
                if not i:
                    continue
                name = i[0]
                if name != 'preptime':
                    ingredient = self.__check_ingredient__(name, recipe)
                    ing_list.append((ingredient[name], i[1]))
                else:
                    prep = i[1]
                    if not prep:
                        prep = 1
            try:
                # TODO Test this
                notes = self.store.recipes[recipe][2]
            except:
                pass
            try:
                # TODO Test this
                yld = self.store.recipes[recipe][3]
            except:
                pass
            #print(ing_list)
            self.store.recipes[recipe] = Recipe(recipe, ing_list, prep, notes, yld)   

class Store:
    def __init__(self, store):
        """ Data and functions pertaining to a specific Store (i.e. restaurant)\n
        Required 'store' is a json db file dict containing:\n
            creds {db_file: (name of the database file), id: (assigned store id number), branch: (full name of the store)}\n
            costs {labor: (current average cost of labor for this store), markup: (percentage to markup recipes for consumers)}\n
            sql_creds {host: (location of database), database: (database name), user: (login user name), pass: (login user password)}\n
            ingredients {ingredients stored in the databse as:\n 
                            'name : (ingredients are always lowercase) {\n
                                name:, \n
                                unit: (measurement i.e. slices of bread or oz of mustard), \n
                                quantity: (how many total in a package, i.e. 24 slices of cheeze or 32 oz of ketchup), \n
                                price: (total price of a package), \n
                                calories: (optional. total calories in a package), \n
                                servings: (optional. total servings in a package), \n
                                notes: (instructions or whatever is needed), \n
                                vendor: (who sells this to the restaurant), \n
                                uuid: (unique item identifier)}}\n
            recipes {recipes stored in the database as:\n
                            Name : (Recipes are always capitalized) {\n
                                name:,\n
                                preptime: (how long this takes to make),\n
                                ingredients: [list of __tuples___: ingredient name, how many (unit of measurement) in recipe (i.e. (cheese, 1))],\n
                                notes: (instructions),\n
                                yld: (how much does this recipe make),\n
                                togo: (packaging associated with recipe)\n
                                is_ingredient: (bool. Recipe used in another rrecipe, i.e Special Sauce Recipe used as special sauce in Burger Recipe),\n
                                uuid: (unique item identifier)}}"""
        self.creds = store['creds']
        self.branch = self.creds['branch']
        self.id = self.creds['id']
        self.db_file = self.creds['db_file']
        self.costs = store['costs']
        global LABOR
        global MARKUP
        LABOR = self.costs['labor']
        MARKUP = self.costs['markup']
        self.sql_creds = store['sql_creds']
        self.ingredients_raw = store['ingredients']
        self.recipes_raw = __decode_json__(store['recipes'])
        if self.ingredients_raw:
            self.ingredients = self.__construct_ingredients__()
        else:
            self.ingredients = self.ingredients_raw
        if self.recipes_raw:
            self.recipes = self.__construct_recipes__()
        else:
            self.recipes = self.recipes_raw
        self.all_items = {**self.ingredients, **self.recipes}
        
    def _json(self):
        db = {}
        db['creds'] = { 'db_file': self.db_file, 'id': self.id, 'branch': self.branch}
        db['costs'] = {'labor': LABOR, 'markup' : MARKUP}
        db['sql_creds'] = self.sql_creds
        db['ingredients'] = __encode_json__(self.ingredients)
        db['recipes'] = __encode_json__(self.recipes)
        return db

    def __construct_ingredients__(self):
        # Recreates ingredients{name : Ingredient (class)} from json data
        ing_dct = __decode_json__(self.ingredients_raw)
        ingredients = {}
        for name, i in ing_dct.items():
            ingredients[name] = Ingredient( i['name'], 
                                            i['unit'], 
                                            i['quantity'], 
                                            i['price'], 
                                            calories=i['calories'], 
                                            servings=i['servings'], 
                                            notes=i['notes'], 
                                            vendor=i['vendor'], 
                                            uuid=i['uuid'])
        return ingredients

    def __construct_recipes__(self):
        # Recreates recipes{name: Recipe (class)} from json data
        recipes = {}
        for name, recipe in self.recipes_raw.items(): # burger, burger dict in recipes
            d = {}
            #print(f'{name=}  {recipe=}  in recipes_dct')
            for k, v in recipe.items():
                #print(f'{k=}  {v=}  in val')
                if k == 'ingredients':
                    inglist = []
                    for ing in recipe[k]:
                        #print(f'{ing[0]=}  {ing[1]=}')
                        inglist.append((self.ingredients[(ing[0])], ing[1]))
                    d[k] = inglist
                else:
                    d[k] = v #d['burger']['name'] = 'burger'
            #for k, v in d.items():
            #    print(f'{k=} : {v=}')
            recipes[name] = Recipe(d['name'], d['ingredients'], d['preptime'], d['notes'], d['yld'], d['togo'], uuid=d['uuid'])
        return recipes 

class Ingredient:
    """Ingredient object containing the following attributes:\n
        name (required lowercase),\n
        unit (how is this ingredient measured, i.e. oz, or slices),\n
        quantity (how many of unit total in a package, i.e. 24oz or 50 slices),\n
        price (total price of package)\n
        optional:\n
            calories (total calories in package),\n
            servings (total servings in a package),\n
            notes (prep instructions, etc.),\n
            vendor (who sells this package),\n
            is_recipe (bool. does this ingredient also exsist in the recipe databse, i.e. Special Sauce),\n
            uuid (unique identifier for each item in database)
    """
    def __init__(self, name, unit, quantity, price, calories=0, servings=1, notes=None, vendor=None, is_recipe=False, uuid=None):
        self.uuid = uuid
        if not self.uuid:
            self.uuid = __generate_id__('i')
        self.name = name.lower()
        self.unit = unit
        self.notes = notes
        self.vendor = vendor
        self.quantity = float(quantity)
        self.price = float(price)
        self.price_per_unit = self.price / self.quantity
        self.calories = round(float(calories),2)
        self.servings = float(servings)
        self.total_calories = self.calories * self.servings
        self.is_recipe = bool(is_recipe)
        try:    
            self.calories_per_unit = round(self.total_calories / self.quantity)
        except ZeroDivisionError:
            self.calories_per_unit = 0
        # Returns only items and not functions in the object
        self.item_dict = {k:v for k, v in self.__dict__.items() if not (k.startswith('__') and k.endswith('__'))}

    def _json(self):
        # return a dict for json output
        d = {}
        d['name'] = self.name
        d['unit'] = self.unit
        d['quantity'] = self.quantity
        d['price'] = self.price
        d['calories'] = self.calories
        d['servings'] = self.servings
        d['notes'] = self.notes
        d['vendor'] = self.vendor
        d['uuid'] = self.uuid
        return d

    def print_ingredient(self):
        # Old printing method
        print(f'{self.name} | {self.quantity} {self.unit} | ${self.price}', end='')
        if self.calories != 0:
            print(f' | calories {self.calories} | servings {self.servings} | calories per {self.unit} {self.calories_per_unit}', end='')
        if self.vendor:
            print(f' | vendor : {self.vendor}', end='')
        if self.notes:
            print(f' | notes : {self.notes}', end='')
        print('')

    def _print_ingredient(self):
        # Dataframe print
        out = {}
        for k, v in self._json().items():
            out[k] = [v]
        out.popitem() # remove the last item (uuid) from the print
        name = out['name']
        df = pd.DataFrame.from_dict(out)
        print(df)

    def query_prices(self):
        # TODO Get prices from vendors
        raise NotImplementedError

class Recipe:
    """Recipe object contianing the folloing attributes:\n
        name (names are requied to be Capitalized),\n
        ingredients (list of tuples containign ingredient name and quantity, i.e. (bread, 2)),\n
        preptime (how long does it take to make, default 1 minute),\n
        optional:\n
            notes (recipe instructions, etc.),\n
            yld (how much does this make, default 1, i.e. 1 Burger)\n
            togo_pack (NotYetImplemented),\n
            is_ingredient (bool. Recipe also used in other recipies, i.e. Special Sauce),\n
            uuid (unique identifier for each item in a databse)\n
    """
    def __init__(self, name, ingredients=[], preptime=1, notes=None, yld=1, togo_pack=None, is_ingredient=False, uuid=None):
        self.uuid = uuid
        if not self.uuid:
            self.uuid = __generate_id__('r')
        self.name = capwords(name)
        self.preptime = float(preptime)
        self.ingredients = ingredients # tuple with Ingredient()[0] and quantity[1]
        self.ingredient_list = []
        self.notes = notes
        self.yld = yld
        self.calories = 0
        self.togo = togo_pack
        self.__display_calories__ = True
        for ingredient in self.ingredients:
            self.ingredient_list.append(ingredient[0].name + ' : ' + str(ingredient[1]) + ' ' + ingredient[0].unit)
            if ingredient[0].calories_per_unit == 0:
                self.__display_calories__ = False # Don't display calories in self.print_recipe() if some ingredients don't have calories
            self.calories += ingredient[0].calories_per_unit * ingredient[1]
        self.is_ingredient = is_ingredient
        self.as_ingredient = None
        if is_ingredient:
            self.as_ingredient = self.create_ingredient()

    def create_ingredient(self):
        ing = self.name.lower()
        print(f'\n~~ Ingredient \'{ing}\' details ~~')
        uid = 'i'+ self.uuid[1:]
        unit = input('Unit of measurement: ')
        yeild = input(f'Recipe yeild ({unit}): ')
        price = Cost(self).total_cost
        calories = 0 # TODO total cals from ingredients
        notes = input('Additional notes: ')
        vendor = '*made in-house'
        #print(f'{uid=}')
        #print(f'{self.uuid=}')
        return Ingredient(ing, unit, yeild, price, calories=calories, notes=notes, vendor=vendor, is_recipe=True, uuid=uid)

    def _food_cost(self):
        # This function has to be called outside the object so it has updated ingredient prices
        cost = 0.0
        for i in self.ingredients:
            #print(i[0].name + ': ' + str(i[1]) + ' * ' + str(i[0].price_per_unit))
            cost += i[1] * i[0].price_per_unit
        return round(cost, 2)

    def print_recipe(self):
        try:
            cost = Cost(self)
            print('\n' + self.name.upper())
            print(self.ingredient_list)
            if self.__display_calories__:
                print('Calories          : ' + str(round(self.calories, 2)))
            print('Food cost         : $' + str(round(self._food_cost(), 2))) # TODO this is inaccurate
            print('Labor cost        : $' + str(round(cost.labor_cost, 2)))
            print('Total cost        : $' + str(round(cost.total_cost, 2)))
            print('Recommended price : $' + str(round(cost.recommended_price, 2)))
        except Exception as e:
            pass

    def _json(self):
        # Creates a dict for json output
        d = {}
        d['name'] = self.name
        d['preptime'] = self.preptime
        # TODO extract ingredient data? or just call name and reconstruct
        ilist = []
        for i in self.ingredients:
            ilist.append(['__tuple__', i[0].name, i[1]])
        d['ingredients'] = ilist
        d['notes'] = self.notes
        d['yld'] = self.yld
        d['togo'] = self.togo
        d['uuid'] = self.uuid
        return d

class Packaging:
    def __init__(self):
        raise NotImplementedError

class Cost:
    """Calculations for recipe LABOR and MARKUP"""
    def __init__(self, recipe):
        global LABOR
        global MARKUP
        self.labor_cost = (recipe.preptime / 60) * LABOR
        self.total_cost = recipe._food_cost() + self.labor_cost
        self.recommended_price = recipe._food_cost() * (MARKUP / 100)
        # TODO yld outputs per recipe

# TODO Update and add to FoodCost
def connect_sql(sql_creds):
    # TODO Flesh this out
    creds = {}
    ings = {}
    recs = {}
    while True:
        # TODO creds should come from store creds and be named after the store id
        print('\n~~~ SERVER CREDENTIALS ~~~\n(type \'done\' to end)')
        creds['host'] = input('Host : ')
        if creds['host'] == 'done':
            return ings, recs
        creds['database'] = input('Database (Resturaunt ID) : ')
        creds['user'] = input('User : ')
        creds['pass'] = input('Pass : ')
        
        # SAMPLE DATA
        creds = sql_creds

        ings = get_sql(creds, 'ingredients')
        recs = get_sql(creds, 'recipes')
        return ings, recs
        #if not ings:
        #    continue
        #else:   
        #    return ings, recs

# TODO Update and add to FoodCoast
def get_sql(creds, table_name):
    count = 0
    err = 0
    outdict = {}
    Table_name = table_name[:-1].capitalize()
    try:
        connection = mysql.connector.connect(host=creds['host'],
        database=creds['database'],
        user=creds['user'],
        password=creds['pass'])

        sql_select_Query = 'SELECT * FROM ' + table_name
        cursor = connection.cursor(dictionary=True)
        cursor.execute(sql_select_Query)
        records = cursor.fetchall()

        for item in records:
            if item['name']:
                try:
                    if table_name == 'ingredients':
                        outdict[item['name']] = Ingredient(item['name'], item['unit'], item['quantity'], item['price'], vendor=item['vendor'], notes=item['notes'])
                    elif table_name == 'recipes':
                        outdict[item['name']] = Recipe(item['name'], item['ingredients'], item['preptime'], item['notes'], item['yld'], item['togo_pack'])
                except:
                    err = err + 1  
                    count = count + 1
    except Error as e:
        print(f'SQL Error: {e}')
        return None
    finally:
        try:
            if (connection.is_connected()):
                connection.close()
                cursor.close()
                #print("MySQL connection is closed")
        except:
            pass
        print(f'\nAdded {count} {table_name}.', end='')
        if err > 0:
            print(f' {err} {Table_name}(s) could not be added due to missing data requirements.')
        else:
            print('')
    return outdict

def __generate_id__(prefix):
    """Creates an unique id for each object, for database storage
    Objects are recreated when decoded from json, so it is necessary to
    create the id only when the opject is first created"""
    ide = uuid.uuid4()
    return prefix + ide.hex

def __decode_json__(d):
    """Creates dict containing proper data types including tuples and objects"""
    outdict = {}
    for k, v in d.items():
        if isinstance(v, str) or isinstance(v, int) or isinstance(v, float):
            outdict[k] = v
        elif v is None:
            outdict[k] = v
        elif isinstance(v, list):
            if len(v) == 3 and v[0] == '__tuple__':
                outdict[k] = (v[1], v[2])
            else:
                temp = []
                for i in v:
                    temp.append(__list_gen__(i))   
                outdict[k] = temp
        elif isinstance(v, dict):
            outdict[k] = __decode_json__(v)
        elif isinstance(v, object) and not (isinstance(v, list) or isinstance(v, tuple)):
            outdict[k] = __decode_json__(v.__dict__)

    return outdict

def __list_gen__(v):
    """Creates a list containing proper data types"""
    if isinstance(v, str) or isinstance(v, int) or isinstance(v, float):
        return v
    elif v is None:
        return v
    elif isinstance(v, list):
        if len(v) == 3 and v[0] == '__tuple__':
                return (v[1], v[2])
        else:
            temp = []
            for i in v:
                temp.append(__list_gen__(i))   
            return temp
    elif isinstance(v, dict):
        return __decode_json__(v)
    elif isinstance(v, object) and not (isinstance(v, list) or isinstance(v, tuple)):
        return __decode_json__(v.__dict__)

def __load_json__(filename, sub=None):
    """Returns the dict from a json file"""
    try:
        with open(filename) as f:
            data = json.load(f)
    except Exception as e:
        sys.exit(f'  Unable to load JSON data. {e}')
    if sub:
        data = data[sub]
    return data

def __encode_json__(dct):
    """Prepares dct for json dump"""
    d = {}
    for k, v in dct.items():
        if isinstance(v, object) and not (isinstance(v, list) or isinstance(v, tuple)):
            # Creates current data from each object when this is called
            d[k] = v._json()
        else:
            d[k] = v
    return d

# Debugging
def print_json(json_dict, keys=None):
    """Debugging tool for printing dicts in a more friendly way"""
    try:
        for k, v in json_dict.items():
            if isinstance(v, dict):
                print_json(v)
            else:
                print(f'{k} : {v}')
    except Exception as e:
        print(f'ERROR: {e}')

if __name__ == '__main__':    
    # args are optional provided command line email and password to skip login requirement
    args = None
    if len(sys.argv) == 3:
        args = sys.argv
    Food_Cost(args)
    